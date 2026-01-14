import os
import tempfile
import zipfile
from pathlib import Path
import base64
import shutil
from datetime import datetime, time, timedelta
import hashlib
import html
import csv
from typing import Any

import geopandas as gpd
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import json
from io import StringIO
import unicodedata
import pyogrio

# Prefer GDAL-backed pyogrio engine to avoid requiring fiona wheels in hosted environments
gpd.options.io_engine = "pyogrio"

# Create rectangles from points and support rotation
from shapely.geometry import box
from shapely.affinity import rotate
import pydeck as pdk

# Optional fiona: use if available for extra drivers (e.g., FileGDB), but do not require it
try:
    import fiona
    from fiona.drvsupport import supported_drivers

    supported_drivers["FileGDB"] = "rw"
    supported_drivers["OpenFileGDB"] = "r"
    HAS_FIONA = True
except Exception:
    HAS_FIONA = False

REFERENCE_DATA_DIR = Path(__file__).parent / "reference_data"
HERO_IMAGE_PATH = Path(__file__).parent / "rwanda_small_map.jpg"
SUPPORTED_REFERENCE_EXTENSIONS = (".xlsx", ".xlsm")
PREVIEW_ROW_COUNT = 20

# Persistent name-memory file (maps equipment_type -> user-chosen filename)
NAME_MEMORY_PATH = Path(__file__).parent / "name_memory.json"


def rerun_app():
    """Trigger a Streamlit rerun across both legacy and new APIs."""
    rerun_callback = getattr(st, "rerun", None)
    if rerun_callback is None:
        rerun_callback = getattr(st, "experimental_rerun", None)
    if rerun_callback is None:
        raise RuntimeError("Unable to rerun Streamlit app: rerun API not available")
    rerun_callback()


def load_name_memory() -> dict:
    """Load the name memory JSON file if present, otherwise return empty dict."""
    try:
        if NAME_MEMORY_PATH.exists():
            with open(NAME_MEMORY_PATH, "r", encoding="utf-8") as fh:
                return json.load(fh)
    except Exception:
        # Corrupt or unreadable file - ignore and start fresh
        return {}
    return {}


def save_name_memory(mapping: dict):
    """Persist the mapping to disk (best-effort)."""
    try:
        with open(NAME_MEMORY_PATH, "w", encoding="utf-8") as fh:
            json.dump(mapping, fh, ensure_ascii=False, indent=2)
    except Exception:
        # Silently ignore failures to avoid disrupting the app UX
        pass


def set_saved_name(equipment_type: str, name: str, memory: dict):
    """Save a single mapping (in-memory and persist to disk)."""
    if not equipment_type or not name:
        return
    memory[equipment_type] = name
    save_name_memory(memory)


# Load memory on startup
name_memory = load_name_memory()

# Persistent UI settings (hero layout etc.)
UI_SETTINGS_PATH = Path(__file__).parent / "ui_settings.json"


def load_ui_settings() -> dict:
    try:
        if UI_SETTINGS_PATH.exists():
            with open(UI_SETTINGS_PATH, "r", encoding="utf-8") as fh:
                return json.load(fh)
    except Exception:
        return {}
    return {}


def save_ui_settings(mapping: dict):
    try:
        with open(UI_SETTINGS_PATH, "w", encoding="utf-8") as fh:
            json.dump(mapping, fh, ensure_ascii=False, indent=2)
    except Exception:
        pass


def load_base64_image(image_path: Path) -> str:
    """Return the base64 representation of an image, or empty string on failure."""
    try:
        with open(image_path, "rb") as fh:
            return base64.b64encode(fh.read()).decode("utf-8")
    except Exception:
        return ""


# load UI settings
ui_settings = load_ui_settings()

# Determine hero height (px) from saved UI settings, fall back to 320
DEFAULT_HERO_HEIGHT = 320
ui_hero_height = int(ui_settings.get("hero_height", DEFAULT_HERO_HEIGHT))
# Determine hero left column percentage (defaults to 35%)
DEFAULT_HERO_LEFT_PCT = 35
ui_hero_left_pct = int(ui_settings.get("hero_left_pct", DEFAULT_HERO_LEFT_PCT))
# Determine hero right column percentage (defaults to 65%)
DEFAULT_HERO_RIGHT_PCT = 65
ui_hero_right_pct = int(ui_settings.get("hero_right_pct", DEFAULT_HERO_RIGHT_PCT))
# Default fixed-left pixel width (used when locking left column in px)
DEFAULT_HERO_LEFT_PX = int(ui_settings.get("hero_left_px", 420))
# Default gradient overlay opacity stops
DEFAULT_HERO_GRADIENT_START = 0.35
DEFAULT_HERO_GRADIENT_END = 0.55
ui_hero_gradient_start = float(ui_settings.get("hero_gradient_start", DEFAULT_HERO_GRADIENT_START))
ui_hero_gradient_end = float(ui_settings.get("hero_gradient_end", DEFAULT_HERO_GRADIENT_END))

# Ensure session state defaults exist so slider changes will produce live preview
if "hero_height_slider" not in st.session_state:
    st.session_state["hero_height_slider"] = ui_settings.get("hero_height", DEFAULT_HERO_HEIGHT)
if "hero_left_pct" not in st.session_state:
    st.session_state["hero_left_pct"] = ui_settings.get("hero_left_pct", DEFAULT_HERO_LEFT_PCT)
if "hero_right_pct" not in st.session_state:
    st.session_state["hero_right_pct"] = ui_settings.get("hero_right_pct", DEFAULT_HERO_RIGHT_PCT)
if "hero_mode" not in st.session_state:
    # modes: 'percent' (default) or 'fixed_left'
    st.session_state["hero_mode"] = ui_settings.get("hero_mode", "percent")
if "hero_left_px" not in st.session_state:
    st.session_state["hero_left_px"] = ui_settings.get("hero_left_px", DEFAULT_HERO_LEFT_PX)
if "hero_gradient_start" not in st.session_state:
    st.session_state["hero_gradient_start"] = ui_settings.get(
        "hero_gradient_start", ui_hero_gradient_start
    )
if "hero_gradient_end" not in st.session_state:
    st.session_state["hero_gradient_end"] = ui_settings.get(
        "hero_gradient_end", ui_hero_gradient_end
    )

# Pre-load hero background image (best-effort) with configurable gradient overlay
hero_bg_data = load_base64_image(HERO_IMAGE_PATH)
hero_gradient_start_used = float(
    min(max(st.session_state.get("hero_gradient_start", ui_hero_gradient_start), 0.0), 1.0)
)
hero_gradient_end_used = float(
    min(max(st.session_state.get("hero_gradient_end", ui_hero_gradient_end), 0.0), 1.0)
)
hero_background_layers = [
    "linear-gradient(135deg, rgba(255, 255, 255, {start:.2f}) 0%, "
    "rgba(248, 250, 252, {end:.2f}) 100%)".format(
        start=hero_gradient_start_used,
        end=hero_gradient_end_used,
    )
]
if hero_bg_data:
    hero_background_layers.append(f"url('data:image/jpeg;base64,{hero_bg_data}')")
hero_background_css = ", ".join(hero_background_layers)


def _reset_stream(stream):
    """Seek to the beginning of a stream if possible."""
    if hasattr(stream, "seek"):
        try:
            stream.seek(0)
        except Exception:
            pass


INVISIBLE_HEADER_CHARS = ["\ufeff", "\u200b", "\u200c", "\u200d", "\ufeff", "\xa0"]
COMPARISON_IGNORED_CHARS = " -_,./()\\"
COMPARISON_TRANSLATION_TABLE = str.maketrans("", "", COMPARISON_IGNORED_CHARS)
MAX_GPKG_NAME_LENGTH = 254


def remove_all_unicode_whitespace(text: str) -> str:
    """
    Remove ONLY non-standard Unicode space-separator characters while
    preserving normal ASCII spaces in the visible text.

    - We still strip Excel artifacts such as:
        U+202F (NARROW NO-BREAK SPACE)
        U+00A0 (NO-BREAK SPACE)  [also handled via INVISIBLE_HEADER_CHARS]
        U+2007, U+2009, U+200A, etc.
    - But we KEEP the regular space character U+0020 so headers look normal.

    This function is safe to use in _clean_column_name() and keeps headers
    readable like:
        "Current Carrying Capacity in Air - Continuous Load (A)"
    instead of:
        "CurrentCarryingCapacityinAir-ContinuousLoad(A)"
    """
    if not isinstance(text, str):
        return text

    cleaned_chars = []
    for ch in text:
        cat = unicodedata.category(ch)
        # Category "Zs" = space separators; we only strip the exotic ones.
        if cat == "Zs" and ch != " ":
            # Drop narrow NBSP, NBSP, thin space, etc.
            continue
        cleaned_chars.append(ch)

    return "".join(cleaned_chars)


def _clean_column_name(name: Any) -> str:
    """
    Clean column names for UI display and GPKG writing.
    This patch improves cleaning by:
    - Removing ALL forms of Unicode whitespace (Zs category)
    - Removing zero-width and BOM characters
    - Collapsing internal spacing
    - Preserving punctuation (important!)
    """

    text = "" if name is None else str(name)

    # Remove ALL unicode whitespace separators (Excel often uses U+202F)
    text = remove_all_unicode_whitespace(text)

    # Remove invisible characters defined in INVISIBLE_HEADER_CHARS
    for ch in INVISIBLE_HEADER_CHARS:
        text = text.replace(ch, "")

    # Collapse multiple visible spaces into a single one
    text = " ".join(text.split())

    return text.strip()


def _strip_comparison_punctuation(text: str) -> str:
    """Remove punctuation and spacing noise for comparison-only matching."""

    if not text:
        return ""
    return text.translate(COMPARISON_TRANSLATION_TABLE)


def normalize_for_compare(name: Any) -> str:
    """
    Normalize a column name for matching by removing:
    - spaces
    - underscores
    - hyphens
    - commas
    - periods
    - parentheses
    - slashes
    - invisible chars
    - converting to lowercase
    """

    if name is None:
        return ""

    text = str(name).lower()

    # remove invisible characters
    for ch in INVISIBLE_HEADER_CHARS:
        text = text.replace(ch, "")

    # collapse whitespace
    text = " ".join(text.split())

    # remove punctuation and comparison noise
    remove_chars = " -_,./()\\"
    text = text.translate(str.maketrans("", "", remove_chars))

    return text.strip()


def normalize_value_for_compare(value: Any) -> str:
    if value is None:
        text = ""
    else:
        try:
            text = "" if pd.isna(value) else str(value)
        except Exception:
            text = str(value)
    text = text.lower()
    for ch in INVISIBLE_HEADER_CHARS:
        text = text.replace(ch, "")
    text = text.replace("_", "").replace("-", "")
    text = " ".join(text.split()).strip()
    return text


def detect_normalized_collisions(series: pd.Series) -> dict[str, set[str]]:
    """
    Return mapping of normalized value -> set of distinct raw values when
    multiple different raw values collapse to the same normalized key.
    """
    collisions: dict[str, set[str]] = {}
    try:
        for value in series.dropna():
            normalized = normalize_value_for_compare(value)
            if not normalized:
                continue
            bucket = collisions.setdefault(normalized, set())
            bucket.add(str(value))
        return {norm: raw_vals for norm, raw_vals in collisions.items() if len(raw_vals) > 1}
    except Exception:
        return {}


def _finalize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame):
        return df
    df.columns = [_clean_column_name(col) for col in df.columns]
    return df


def _apply_global_forward_fill(df: pd.DataFrame) -> pd.DataFrame:
    """Replace empty strings with NA and forward-fill the entire DataFrame."""
    if isinstance(df, pd.DataFrame):
        df = df.replace("", pd.NA)
        df = df.ffill()
    return df


def _is_effectively_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    try:
        return pd.isna(value)
    except Exception:
        return False


def _stringify_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except Exception:
            return value.decode("utf-8", errors="replace")
    if isinstance(value, (list, dict, set, tuple)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    if isinstance(value, (datetime, timedelta)):
        return value.isoformat()
    return str(value)


def derive_layer_name_from_filename(file_name: str) -> str:
    """Return a GeoPackage-safe layer name derived from an uploaded filename."""
    try:
        stem = Path(file_name).stem
    except Exception:
        stem = ""

    cleaned = stem.strip()
    if not cleaned:
        cleaned = "layer"

    if len(cleaned) > MAX_GPKG_NAME_LENGTH:
        cleaned = cleaned[:MAX_GPKG_NAME_LENGTH]

    return cleaned


def ensure_valid_gpkg_dtypes(series: pd.Series) -> pd.Series:
    """Coerce a pandas Series into a GPKG-safe dtype."""
    if not isinstance(series, pd.Series):
        return series

    result = series.copy()

    if pd.api.types.is_datetime64tz_dtype(result):
        result = result.dt.tz_localize(None)
    elif pd.api.types.is_datetime64_any_dtype(result):
        # Already naive datetime - leave as-is
        pass
    elif pd.api.types.is_timedelta64_dtype(result):
        result = result.astype(str)

    # If we see complex/mixed objects, convert everything to string
    if pd.api.types.is_object_dtype(result) or any(
        isinstance(v, (list, dict, set, tuple)) for v in result.dropna().head(5)
    ):
        result = result.apply(_stringify_value)

    if pd.api.types.is_numeric_dtype(result):
        # Ensure a concrete numeric dtype (avoids weird pandas extension types)
        result = result.astype("float64" if pd.api.types.is_float_dtype(result) else result.dtype)

    return result


def _truncate_column_name(name: str, used: dict) -> str:
    base = name if len(name) <= MAX_GPKG_NAME_LENGTH else name[:MAX_GPKG_NAME_LENGTH]
    candidate = base
    counter = 1
    while candidate in used:
        suffix = f"_{counter}"
        limit = MAX_GPKG_NAME_LENGTH - len(suffix)
        candidate = (base[:limit] if len(base) > limit else base) + suffix
        counter += 1
    used[candidate] = True
    return candidate


def parse_pasted_tabular_text(text: str) -> pd.DataFrame:
    """Parse raw pasted TSV/CSV text into a DataFrame."""
    cleaned = text.replace("\r", "\n")
    cleaned = cleaned.replace("\n\n", "\n")
    for ch in INVISIBLE_HEADER_CHARS:
        cleaned = cleaned.replace(ch, "")
    cleaned = cleaned.strip()
    if not cleaned:
        return pd.DataFrame()

    sample = cleaned[:10000]
    delimiters = ["\t", ",", ";", "|", "\u0001"]
    sep = "\t" if "\t" in sample else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=delimiters)
        sep = dialect.delimiter
    except csv.Error:
        if ";" in sample and sep not in ("\t", ","):
            sep = ";"

    df = pd.read_csv(
        StringIO(cleaned),
        sep=sep,
        engine="python",
        dtype=str,
        keep_default_na=False,
        na_filter=False,
    )
    df = _apply_global_forward_fill(df)
    return _finalize_dataframe_columns(df)


def read_tabular_data(source):
    """
    Load a CSV/Excel file while preserving headers and raw text as much as possible.

    NOTE: We avoid deprecated/removed pandas kwargs (like mangle_dupe_cols)
    so this stays compatible with pandas 2.x.
    """
    if isinstance(source, (str, Path)):
        suffix = Path(source).suffix.lower()
    else:
        suffix = Path(source.name).suffix.lower()

    # Common CSV options - autodetect separator via sep=None with python engine
    csv_kwargs = {
        "dtype": str,
        "keep_default_na": False,
        "na_filter": False,
        "sep": None,
        "engine": "python",
    }

    if suffix == ".csv":
        encodings = ("utf-8-sig", "utf-16", "utf-8", "latin-1")
        for encoding in encodings:
            _reset_stream(source)
            try:
                df = pd.read_csv(source, encoding=encoding, **csv_kwargs)
                df = _apply_global_forward_fill(df)
                return _finalize_dataframe_columns(df)
            except UnicodeDecodeError:
                continue

        # Last resort with replacement characters
        _reset_stream(source)
        df = pd.read_csv(
            source,
            encoding="utf-8",
            errors="replace",
            **csv_kwargs,
        )
        df = _apply_global_forward_fill(df)
        return _finalize_dataframe_columns(df)

    if suffix in {".xlsx", ".xlsm", ".xls"}:
        _reset_stream(source)
        df = pd.read_excel(
            source,
            dtype=str,
            na_filter=False,
            keep_default_na=False,
        )
        df = _apply_global_forward_fill(df)
        return _finalize_dataframe_columns(df)

    raise ValueError(f"Unsupported file type: {suffix}")


def _safe_extract_zip(zip_file: zipfile.ZipFile, dest_dir: str):
    """
    Safely extract a ZIP file, rejecting absolute paths and traversal attempts.
    Returns a list of extracted file paths.
    """
    dest_path = Path(dest_dir).resolve()
    extracted = []

    for member in zip_file.infolist():
        member_path = Path(member.filename)

        if member_path.is_absolute():
            raise ValueError(f"Blocked absolute path in ZIP: {member.filename}")

        resolved = (dest_path / member_path).resolve()
        if dest_path not in resolved.parents and dest_path != resolved:
            raise ValueError(f"Blocked path traversal in ZIP: {member.filename}")

        if member.is_dir():
            resolved.mkdir(parents=True, exist_ok=True)
            continue

        resolved.parent.mkdir(parents=True, exist_ok=True)
        with zip_file.open(member, "r") as source, open(resolved, "wb") as target:
            shutil.copyfileobj(source, target)
        extracted.append(str(resolved))

    return extracted


def clean_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Remove only rows that are entirely empty while preserving ALL columns."""
    try:
        if not isinstance(df, pd.DataFrame):
            return df
        if df.empty:
            return df

        mask = df.apply(lambda col: col.map(_is_effectively_empty))
        cleaned = df.loc[~mask.all(axis=1)].copy()
        cleaned.columns = list(df.columns)
        cleaned = _apply_global_forward_fill(cleaned)
        return cleaned
    except Exception:
        return df


def get_reference_workbooks():
    """Return mapping of workbook label -> path for bundled Excel files."""
    if not REFERENCE_DATA_DIR.exists():
        return {}

    workbooks = {}
    for workbook in sorted(
        p
        for p in REFERENCE_DATA_DIR.rglob("*")
        if p.is_file() and p.suffix.lower() in SUPPORTED_REFERENCE_EXTENSIONS
    ):
        label = workbook.relative_to(REFERENCE_DATA_DIR).as_posix()
        workbooks[label] = workbook

    return workbooks


def get_sheet_names(workbook_path: Path):
    """Return available sheet names for the selected workbook."""
    try:
        excel_file = pd.ExcelFile(workbook_path)
        return excel_file.sheet_names
    except Exception:
        return []


def describe_reference_sheet(workbook_path: Path, sheet_name: str):
    """Return metadata describing the requested worksheet."""
    wb = None
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        worksheet = wb[sheet_name]
        header_values = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )
        headers = [value for value in header_values if value is not None]
        row_count = max(worksheet.max_row - (1 if headers else 0), 0)
        column_count = worksheet.max_column
        return {
            "rows": row_count,
            "columns": column_count,
            "headers": headers,
        }
    except Exception:
        return None
    finally:
        if wb is not None:
            wb.close()


def load_reference_preview(workbook_path: Path, sheet_name: str, max_rows: int = PREVIEW_ROW_COUNT):
    """Return a lightweight preview of the sheet for UI display."""
    try:
        preview = pd.read_excel(
            workbook_path,
            sheet_name=sheet_name,
            nrows=max_rows,
        )
        preview = _apply_global_forward_fill(preview)
        return preview
    except Exception:
        return pd.DataFrame()


st.set_page_config(
    page_title="Clean GPKG Attribute Filler",
    page_icon="gpkg",
    layout="wide",
)

# Use session-state values (if present) so slider changes show live preview
hero_height_used = int(st.session_state.get("hero_height_slider", ui_hero_height))
hero_left_pct_used = st.session_state.get("hero_left_pct", ui_hero_left_pct)
hero_right_pct_used = st.session_state.get("hero_right_pct", ui_hero_right_pct)
# Determine flex rules depending on mode
hero_mode_used = st.session_state.get("hero_mode", "percent")
if hero_mode_used == "fixed_left":
    # left fixed in px, right fills remaining space
    left_flex_css = "0 0 " + str(int(st.session_state.get("hero_left_px", DEFAULT_HERO_LEFT_PX))) + "px"
    right_flex_css = "1 1 auto"
else:
    left_flex_css = "0 0 " + str(int(hero_left_pct_used)) + "%"
    right_flex_css = "0 0 " + str(int(hero_right_pct_used)) + "%"

st.markdown("""
    <style>
    * {
        margin: 0;
        padding: 0;
        box-sizing: border-box;
    }
    .stApp {
        font-family: 'Segoe UI', system-ui, -apple-system, BlinkMacSystemFont, 'Helvetica Neue', sans-serif;
        background: #f4f6f9;
    }
    .main {
        padding: 0 !important;
    }
    .main > div {
        width: 100% !important;
        max-width: 100% !important;
        padding: 0 !important;
        margin: 0 !important;
    }
    
    /* Aggressively override all Streamlit width constraints */
    main .block-container {
        width: 100% !important;
        max-width: 100% !important;
        padding: 0 !important;
        margin: 0 !important;
    }
    section[data-testid="stSidebar"] + div {
        width: 100% !important;
        max-width: 100% !important;
    }
    
    /* Hero Section - Full Width */
    .hero-container {
        display: flex;
        width: 100%;
        max-width: 100% !important;
        min-height: """ + str(hero_height_used) + """px;
        margin: 0 !important;
        padding: 0 !important;
        margin-bottom: 0;
        box-shadow: 0 8px 20px rgba(13, 71, 161, 0.15);
        border-radius: 0 !important;
        overflow: hidden;
    }
    
    /* Hero Left Column - Blue Branding */
    .hero-left {
        flex: """ + left_flex_css + """;
        background: linear-gradient(135deg, #0d47a1 0%, #1565c0 100%);
        color: #ffffff;
        padding: 3rem 2.5rem;
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: flex-start;
    }
    .hero-left h2 {
        font-size: 2.2rem;
        font-weight: 700;
        margin-bottom: 1.5rem;
        letter-spacing: -0.8px;
    }
    .hero-left .tagline {
        font-size: 1rem;
        font-weight: 500;
        color: #bbdefb;
        margin-bottom: 1rem;
        line-height: 1.5;
    }
    .hero-left .byline {
        font-size: 0.9rem;
        color: #90caf9;
        font-style: italic;
        margin-top: 2rem;
        padding-top: 1.5rem;
        border-top: 1px solid rgba(255, 255, 255, 0.2);
    }
    
    /* Hero Right Column - Product Title + Background */
    .hero-right {
        flex: """ + right_flex_css + """;
        background-image: """ + hero_background_css + """;
        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
        padding: 3rem 2.5rem;
        text-align: center;
        position: relative;
    }
    .hero-right::before {
        content: "";
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background: none;
        z-index: 1;
    }
    .hero-right h1,
    .hero-right .subtitle {
        position: relative;
        z-index: 2;
    }
    .hero-right h1 {
        font-size: 2.4rem;
        font-weight: 700;
        color: #0d47a1;
        line-height: 1.3;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
        margin: 0;
        letter-spacing: -0.5px;
    }
    .hero-right .subtitle {
        font-size: 1rem;
        color: #1565c0;
        margin-top: 1rem;
        font-weight: 500;
    }
    
    /* Content Wrapper - Full width, center child elements */
    .content-wrapper {
        width: 100% !important;
        max-width: 100% !important;
        padding: 2rem !important;
        margin: 0 !important;
        display: flex;
        flex-direction: column;
        align-items: center;
    }
    
    .content-wrapper > * {
        width: 100%;
        max-width: 980px;
        margin-left: auto;
        margin-right: auto;
    }
    
    /* Section Box - Main workflow containers */
    .section-box {
        background: #ffffff;
        padding: 2rem;
        border-radius: 12px;
        box-shadow: 0 2px 8px rgba(15, 23, 42, 0.08);
        border-left: 4px solid #2a5298;
        margin-bottom: 2rem;
    }
    .section-box.alt {
        border-left-color: #5a67d8;
    }
    .section-box.tertiary {
        border-left-color: #3b82f6;
    }
    
    .section-title {
        font-size: 1.4rem;
        font-weight: 600;
        color: #1f2a37;
        margin-bottom: 1.2rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    .section-title::before {
        content: "";
        display: inline-block;
        width: 4px;
        height: 1.4rem;
        background: #2a5298;
        border-radius: 2px;
    }
    
    .section-subtext {
        color: #4b5563;
        margin-bottom: 1.5rem;
        font-size: 0.95rem;
        line-height: 1.5;
    }
    
    .stFileUploader {
        border-radius: 12px !important;
    }
    .stFileUploader > div {
        border-radius: 12px !important;
        border: 2px dashed #3b82f6 !important;
        padding: 1.5rem !important;
        background: linear-gradient(135deg, rgba(59, 130, 246, 0.05) 0%, rgba(79, 172, 254, 0.02) 100%) !important;
    }
    
    .stTextInput > div > div,
    .stSelectbox > div > div,
    .stDataEditor > div {
        border-radius: 8px !important;
        border: 1px solid #e5e7eb !important;
    }
    .stTextInput > div > div:focus-within,
    .stSelectbox > div > div:focus-within {
        border-color: #2a5298 !important;
        box-shadow: 0 0 0 3px rgba(42, 82, 152, 0.1) !important;
    }
    
    .stRadio > label {
        font-weight: 500;
        color: #1f2a37;
    }
    .stRadio > div {
        gap: 1rem;
    }
    
    .stButton button {
        font-weight: 600;
        padding: 0.75rem 1.5rem !important;
        border-radius: 8px !important;
        background: linear-gradient(135deg, #2a5298 0%, #3b82f6 100%) !important;
        color: #ffffff !important;
        border: none !important;
        transition: all 0.3s ease;
    }
    .stButton button:hover {
        box-shadow: 0 4px 12px rgba(42, 82, 152, 0.3);
        transform: translateY(-2px);
    }
    
    footer {
        visibility: hidden;
    }
    .custom-footer {
        text-align: center;
        padding: 2rem 0 1rem;
        color: #6b7280;
        font-size: 0.9rem;
        border-top: 1px solid #e5e7eb;
        margin-top: 3rem;
    }
    
    @media (max-width: 768px) {
        .hero-container {
            flex-direction: column;
            min-height: auto;
        }
        .hero-left,
        .hero-right {
            flex: 0 0 100%;
        }
        .hero-left h2 {
            font-size: 1.8rem;
        }
        .hero-right h1 {
            font-size: 1.8rem;
        }
        .section-box {
            padding: 1.5rem;
        }
        .content-wrapper {
            padding: 1rem;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
<div class="hero-container">
    <div class="hero-left">
        <h2>GeoData Fusion</h2>
        <div class="tagline">
            <strong>Smart Attribute Mapping</strong><br>
            Harmonize GeoPackage data with precision
        </div>
        <div class="byline">
            Built by Eng. IRANZI Prince Jean Claude<br>
            For engineers, by engineers.
        </div>
    </div>
    <div class="hero-right">
        <h1>Substations and Power Plants GIS Modelling</h1>
        <div class="subtitle">Professional geospatial data management for Rwanda's infrastructure</div>
    </div>
</div>
""",
    unsafe_allow_html=True,
)

# ---------------------- QUOTE OF THE DAY ---------------------------------
QOTD_PATH = Path(__file__).parent / "quotes.json"
QOTD_REFRESH_TIME = time(6, 0)

DEFAULT_QOTD_QUOTES = [
    {"text": "Measure twice, map once; precision makes spatial insight powerful.", "author": "Surveyor's Axiom"},
    {"text": "Every coordinate tells a story waiting for an engineer to interpret it.", "author": "GeoSystems Lead"},
    {"text": "Accurate data layers are the scaffolding of resilient infrastructure.", "author": "Civil GIS Collective"},
    {"text": "An engineer armed with GIS can turn raw terrain into informed design.", "author": "Spatial Planning Guild"},
    {"text": "Maps are the interface between imagination and construction.", "author": "Ada Augusta"},
    {"text": "When latitude meets logic, breakthroughs follow.", "author": "Control Point Studio"},
    {"text": "Great grids make great cities.", "author": "Urban Network Initiative"},
    {"text": "Never trust a map you didn't debug yourself.", "author": "Field Engineer's Rule"},
    {"text": "Topology errors are whispers that something in the field needs your attention.", "author": "GIS QA Team"},
    {"text": "Scale models fade, but geospatial models evolve with every dataset.", "author": "Digital Twin Lab"},
    {"text": "Precision engineering is a love letter to the future.", "author": "Structures Atelier"},
    {"text": "Terrain is the silent stakeholder in every infrastructure project.", "author": "Hydrology Partners"},
    {"text": "Buffer your assumptions like you buffer your geometries - generously.", "author": "Spatial Analyst Humor"},
    {"text": "Innovation happens where satellite imagery meets stubborn curiosity.", "author": "Orbital Cartography Group"},
    {"text": "The shortest path algorithms teach us: constraints reveal elegance.", "author": "Graph Theory Circle"},
    {"text": "Coordinate systems are the grammar of geographic storytelling.", "author": "Projection Society"},
    {"text": "Clean schemas save muddy boots.", "author": "Field Data Core"},
    {"text": "A resilient grid is engineered twice: once on-site, once on-screen.", "author": "Power Systems Studio"},
    {"text": "Contours are the fingerprints of the earth.", "author": "Topographic Collective"},
    {"text": "Metadata is the engineering diary your future self will thank you for.", "author": "Documentation League"},
    {"text": "In GIS, accuracy is kindness to the crews who follow your plans.", "author": "Pipeline Cartographer"},
    {"text": "Let data drive decisions, but let engineers drive the data.", "author": "Systems Integration Forum"},
    {"text": "Spatial joins turn isolated facts into operational knowledge.", "author": "GeoAnalytics Lab"},
    {"text": "A well-designed attribute table is as vital as a well-cured concrete pour.", "author": "Structural Data Guild"},
    {"text": "Routing fiber or roads, the map cares not - only the engineer's intent matters.", "author": "Infrastructure Weavers"},
    {"text": "Use elevation to your advantage; gravity is the oldest project partner.", "author": "Hydraulic Insights"},
    {"text": "Reliable basemaps are quiet enablers of heroic field days.", "author": "Remote Sensing Crew"},
    {"text": "Quality control in GIS is the compass that keeps projects true north.", "author": "Survey Integrity Team"},
    {"text": "Layer transparency teaches us that clarity often lives in overlap.", "author": "Visualization Studio"},
    {"text": "Engineers who map well build well.", "author": "Site Readiness Council"},
    {"text": "Voltage without vision is noise; vision with voltage powers nations.", "author": "Grid Architects"},
    {"text": "A disciplined circuit diagram is a promise of reliable light.", "author": "Powerline Atelier"},
    {"text": "Great mapping is invisible; people only notice when it is missing.", "author": "Cartography Collective"},
    {"text": "Survey stakes may move, but truth in data should not.", "author": "Field Integrity Corps"},
    {"text": "Engineers turn constraints into catalysts for excellence.", "author": "Design Performance Lab"},
    {"text": "Every raster pixel is a sensor whispering about the earth.", "author": "Imagery Insights"},
    {"text": "Switchyards reward patience - the neatest diagrams prevent the loudest faults.", "author": "Substation Guild"},
    {"text": "Model the future as carefully as you document the past.", "author": "Heritage Engineers"},
    {"text": "A tidy attribute table shortens site visits more than any shortcut road.", "author": "Logistics Cartographers"},
    {"text": "Precision thrives where curiosity meets calibration.", "author": "Metrology Circle"},
    {"text": "Triangulate problems from multiple datasets before they triangulate you.", "author": "Spatial Troubleshooters"},
    {"text": "Sustainable grids start with engineers who listen to the landscape.", "author": "Green Circuit Alliance"},
    {"text": "A GIS without metadata is a switchboard without labels.", "author": "Operations Whisper"},
    {"text": "Let excellence be the standard unit in every engineering drawing.", "author": "Drafting Virtuosi"},
    {"text": "Electrical safety begins with accurate schematics and ends with disciplined crews.", "author": "Protection Relay Team"},
    {"text": "Contours, cables, and code all obey the same rule: clarity first.", "author": "Interdisciplinary Forum"},
    {"text": "When you blend spatial science with empathy, communities flourish.", "author": "Human-Centered GIS"},
    {"text": "Testing assumptions is cheaper than rebuilding substations.", "author": "Reliability Taskforce"},
    {"text": "Bring excellence to mundane tasks; tomorrow's breakthroughs stand on them.", "author": "Continuous Improvement Lab"},
    {"text": "Map the risks before they map you.", "author": "Resilience Cartography"},
]

if not QOTD_PATH.exists():
    try:
        with open(QOTD_PATH, "w", encoding="utf-8") as fh:
            json.dump(DEFAULT_QOTD_QUOTES, fh, ensure_ascii=False, indent=2)
    except Exception:
        pass

loaded_quotes = []
try:
    with open(QOTD_PATH, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    if isinstance(data, list):
        for entry in data:
            text_val = entry.get("text")
            author_val = entry.get("author")
            if text_val and author_val:
                loaded_quotes.append({"text": str(text_val), "author": str(author_val)})
except Exception:
    loaded_quotes = []

quote_pool = loaded_quotes if loaded_quotes else DEFAULT_QOTD_QUOTES
now = datetime.now()
quote_cycle_date = now.date()
if now.time() < QOTD_REFRESH_TIME:
    quote_cycle_date = (now - timedelta(days=1)).date()
quote_cycle_key = quote_cycle_date.isoformat()

if st.session_state.get("qotd_cycle_key") != quote_cycle_key:
    digest = hashlib.sha256(quote_cycle_key.encode("utf-8")).hexdigest()
    quote_index = int(digest, 16) % len(quote_pool)
    st.session_state["qotd_cycle_key"] = quote_cycle_key
    st.session_state["qotd_quote"] = quote_pool[quote_index]

quote_today = st.session_state.get("qotd_quote", quote_pool[0])
quote_text = html.escape(str(quote_today.get("text", "")))
quote_author = html.escape(str(quote_today.get("author", "")))

st.markdown(
    """
    <style>
    .qotd-box {
        background: linear-gradient(135deg, #eef4ff 0%, #ffffff 100%);
        border-radius: 12px;
        padding: 1.5rem;
        text-align: center;
        margin: 1.5rem 0 0.5rem 0;
        box-shadow: 0 20px 35px rgba(15, 23, 42, 0.08);
        border: 1px solid rgba(148, 163, 184, 0.25);
    }
    .qotd-text {
        font-style: italic;
        font-size: 1.15rem;
        color: #1f2937;
        margin-bottom: 0.4rem;
    }
    .qotd-author {
        font-size: 0.95rem;
        color: #475569;
        letter-spacing: 0.02em;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    f"""
    <div class="qotd-box">
        <div class="qotd-text">"{quote_text}"</div>
        <div class="qotd-author">- {quote_author}</div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ---------------- UI SETTINGS EXPANDER -----------------
with st.expander("UI Settings", expanded=False):
    try:
        hero_height_new = st.slider(
            "Hero height (px)",
            min_value=200,
            max_value=800,
            value=st.session_state.get("hero_height_slider", ui_hero_height),
            step=10,
            key="hero_height_slider",
        )

        st.markdown("**Hero background gradient overlay**")
        hero_gradient_start_new = st.slider(
            "Gradient start opacity (0 = transparent, 1 = solid)",
            min_value=0.0,
            max_value=1.0,
            value=float(st.session_state.get("hero_gradient_start", hero_gradient_start_used)),
            step=0.05,
            key="hero_gradient_start",
        )
        hero_gradient_end_new = st.slider(
            "Gradient end opacity (0 = transparent, 1 = solid)",
            min_value=0.0,
            max_value=1.0,
            value=float(st.session_state.get("hero_gradient_end", hero_gradient_end_used)),
            step=0.05,
            key="hero_gradient_end",
        )

        st.markdown("**Hero sizing mode**")
        hero_mode_new = st.radio(
            "Choose how the hero columns size:",
            ("percent", "fixed_left"),
            index=0 if st.session_state.get("hero_mode", "percent") == "percent" else 1,
            key="hero_mode",
        )

        if st.session_state.get("hero_mode", "percent") == "percent":
            hero_left_new = st.number_input(
                "Hero left column width (%) - enter any integer (no limit)",
                value=int(st.session_state.get("hero_left_pct", ui_hero_left_pct)),
                step=1,
                key="hero_left_pct",
            )

            hero_right_new = st.number_input(
                "Hero right column width (%) - enter any integer (no limit)",
                value=int(st.session_state.get("hero_right_pct", ui_hero_right_pct)),
                step=1,
                key="hero_right_pct",
            )
        else:
            hero_left_px_new = st.number_input(
                "Hero left column width (px)",
                value=int(st.session_state.get("hero_left_px", DEFAULT_HERO_LEFT_PX)),
                step=1,
                key="hero_left_px",
            )

        st.markdown("*Live preview updates as you change values. Click Save to persist.*")

        if st.button("Save UI settings", key="save_ui_settings_btn"):
            ui_settings["hero_height"] = int(st.session_state.get("hero_height_slider", hero_height_new))
            ui_settings["hero_mode"] = st.session_state.get("hero_mode", "percent")
            ui_settings["hero_gradient_start"] = float(
                st.session_state.get("hero_gradient_start", hero_gradient_start_new)
            )
            ui_settings["hero_gradient_end"] = float(
                st.session_state.get("hero_gradient_end", hero_gradient_end_new)
            )
            if ui_settings["hero_mode"] == "percent":
                ui_settings["hero_left_pct"] = int(st.session_state.get("hero_left_pct", hero_left_new))
                ui_settings["hero_right_pct"] = int(st.session_state.get("hero_right_pct", hero_right_new))
                ui_settings.pop("hero_left_px", None)
            else:
                ui_settings["hero_left_px"] = int(st.session_state.get("hero_left_px", hero_left_px_new))
                ui_settings.pop("hero_left_pct", None)
                ui_settings.pop("hero_right_pct", None)
            save_ui_settings(ui_settings)
            st.success("Saved UI settings")
            rerun_app()
        if st.button("Reset to defaults", key="reset_ui_settings_btn"):
            ui_settings.pop("hero_height", None)
            ui_settings.pop("hero_left_pct", None)
            ui_settings.pop("hero_right_pct", None)
            ui_settings.pop("hero_mode", None)
            ui_settings.pop("hero_left_px", None)
            ui_settings.pop("hero_gradient_start", None)
            ui_settings.pop("hero_gradient_end", None)
            save_ui_settings(ui_settings)
            st.session_state["hero_height_slider"] = DEFAULT_HERO_HEIGHT
            st.session_state["hero_left_pct"] = DEFAULT_HERO_LEFT_PCT
            st.session_state["hero_right_pct"] = DEFAULT_HERO_RIGHT_PCT
            st.session_state["hero_mode"] = "percent"
            st.session_state["hero_left_px"] = DEFAULT_HERO_LEFT_PX
            st.session_state["hero_gradient_start"] = DEFAULT_HERO_GRADIENT_START
            st.session_state["hero_gradient_end"] = DEFAULT_HERO_GRADIENT_END
            st.success("Reset UI settings to defaults")
            rerun_app()
    except Exception:
        pass

# ---------------------- SINGLE FILE MERGE WORKFLOW ------------------------
reference_workbooks = get_reference_workbooks()

with st.container():
    st.markdown('<div class="section-box">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Single File Upload</div>', unsafe_allow_html=True)
    st.markdown(
        "<p class='section-subtext'>Upload your GeoPackage and select how attribute data should be provided.</p>",
        unsafe_allow_html=True,
    )
    gpkg_files = st.file_uploader(
        "Upload GeoPackage (.gpkg)",
        type=["gpkg"],
        key="single_gpkg",
        accept_multiple_files=True,
    )

    data_source = st.radio(
        "Attribute data source",
        (
            "Upload CSV/Excel file",
            "Use stored reference workbook",
            "Paste data directly",
        ),
        key="data_source_choice",
    )

    uploaded_data_file = None
    reference_sheet = None
    reference_path = None
    workbook_label = None
    shared_df = None
    shared_source_label = None

    if data_source == "Paste data directly":
        st.info(
            "Paste raw tabular data inside each GeoPackage block below. Each dataset has its own paste area and editable preview."
        )

    if data_source == "Upload CSV/Excel file":
        uploaded_data_file = st.file_uploader(
            "Upload Data File (CSV or Excel)",
            type=["csv", "xlsx"],
            key="data_file_uploader",
        )
        if uploaded_data_file is not None:
            try:
                shared_df = read_tabular_data(uploaded_data_file)
                shared_df = clean_empty_rows(shared_df)
                shared_source_label = uploaded_data_file.name
                st.success("Data loaded.")
            except Exception as exc:
                st.error(f"Unable to read uploaded data file: {exc}")
    elif data_source == "Use stored reference workbook":
        if not reference_workbooks:
            st.info(
                "No reference workbooks found in `reference_data`. Add an Excel file to that folder to use this option."
            )
        else:
            workbook_label = st.selectbox(
                "Select stored workbook",
                list(reference_workbooks.keys()),
                key="reference_workbook_select",
            )
            reference_path = reference_workbooks.get(workbook_label)
            sheet_names = get_sheet_names(reference_path) if reference_path else []
            if sheet_names:
                reference_sheet = st.selectbox(
                    "Select worksheet",
                    sheet_names,
                    key="reference_sheet_select",
                )
                if reference_path and reference_sheet:
                    try:
                        shared_df = pd.read_excel(reference_path, sheet_name=reference_sheet)
                        shared_df = _apply_global_forward_fill(shared_df)
                        shared_df = clean_empty_rows(shared_df)
                        shared_source_label = f"{workbook_label} - sheet: {reference_sheet}"
                        st.success(
                            f"Reference workbook loaded ({workbook_label} - sheet: {reference_sheet})"
                        )
                    except Exception as exc:
                        st.error(f"Unable to read selected workbook: {exc}")
            else:
                st.warning("Unable to read sheet names from the selected workbook.")
    st.markdown('</div>', unsafe_allow_html=True)

with st.container():
    st.markdown('<div class="section-box alt">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Output Filename</div>', unsafe_allow_html=True)
    st.markdown(
        "<p class='section-subtext'>Customize how the updated GeoPackage will be saved. The same name is used for the output layer.</p>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p>Each uploaded GeoPackage now has its own naming input, paste block, and merge controls in the sections below.</p>",
        unsafe_allow_html=True,
    )
    st.markdown('</div>', unsafe_allow_html=True)


def sanitize_gdf_for_gpkg(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Prepare a GeoDataFrame for writing to GeoPackage by:
    - Converting unsupported types (object, datetime64 with tz) to GPKG-safe values.
    - Truncating column names to 254 characters (GPKG limit).
    - PRESERVING empty columns (no dropping all-NaN fields).
    """
    gdf_copy = gdf.copy()
    geometry_name = gdf_copy.geometry.name if hasattr(gdf_copy, "geometry") else None

    used_names = {}
    normalized_used = {}

    new_columns = []
    for col in gdf_copy.columns:
        if col == geometry_name:
            new_columns.append(col)
            continue

        clean = _clean_column_name(col)
        normalized = normalize_for_compare(clean)

        if normalized in normalized_used:
            base = clean
            counter = 1
            candidate = f"{base}_{counter}"
            while normalize_for_compare(candidate) in normalized_used:
                counter += 1
                candidate = f"{base}_{counter}"
            clean = candidate

        normalized_used[normalize_for_compare(clean)] = True
        used_names[clean] = True
        new_columns.append(clean)
    gdf_copy.columns = new_columns

    for col in gdf_copy.columns:
        if col == geometry_name:
            continue
        series = ensure_valid_gpkg_dtypes(gdf_copy[col])
        mask = pd.isna(series)
        if mask.any():
            series = series.astype(object)
            series[mask] = None
        gdf_copy[col] = series

    return gdf_copy


def merge_without_duplicates(
    gdf: gpd.GeoDataFrame,
    df: pd.DataFrame,
    left_key: str,
    right_key: str,
) -> gpd.GeoDataFrame:
    """Join df onto gdf with Excel values overwriting non-empty GPKG values."""

    base_gdf = gdf.copy()
    incoming_df = df.copy()

    original_incoming_columns = list(incoming_df.columns)
    incoming_df = _finalize_dataframe_columns(incoming_df)
    cleaned_incoming_columns = list(incoming_df.columns)
    column_lookup = {
        original: cleaned
        for original, cleaned in zip(original_incoming_columns, cleaned_incoming_columns)
    }

    incoming_right_key = column_lookup.get(right_key, _clean_column_name(right_key))
    if incoming_right_key not in incoming_df.columns:
        incoming_right_key = right_key

    # Detect collisions where different raw values collapse to the same normalized key
    collision_messages = []
    left_collisions = detect_normalized_collisions(base_gdf[left_key])
    right_collisions = detect_normalized_collisions(incoming_df[incoming_right_key])
    if left_collisions:
        examples = "; ".join(
            ", ".join(sorted(list(values))[:3]) for values in left_collisions.values()
        )
        collision_messages.append(
            f"GeoPackage join field '{left_key}' has values that normalize to the same key ({examples})."
        )
    if right_collisions:
        examples = "; ".join(
            ", ".join(sorted(list(values))[:3]) for values in right_collisions.values()
        )
        collision_messages.append(
            f"Tabular join field '{right_key}' has values that normalize to the same key ({examples})."
        )
    if collision_messages:
        raise ValueError(
            "Normalized join key collisions detected. "
            + " ".join(collision_messages)
            + " Clean the join fields so each normalized key is unique before merging."
        )

    geometry_name = base_gdf.geometry.name if hasattr(base_gdf, "geometry") else None

    # Normalize incoming columns to match existing GeoPackage columns
    gpkg_norm = {
        normalize_for_compare(col): col
        for col in base_gdf.columns
        if col != geometry_name
    }

    normalized_matches: dict[str, str] = {}
    for col in incoming_df.columns:
        if col == incoming_right_key:
            continue
        norm = normalize_for_compare(col)
        if norm in gpkg_norm:
            normalized_matches[col] = gpkg_norm[norm]

    norm_key = "_norm_key"
    counter = 1
    while norm_key in base_gdf.columns or norm_key in incoming_df.columns:
        norm_key = f"_norm_key_{counter}"
        counter += 1

    base_gdf[norm_key] = base_gdf[left_key].apply(normalize_value_for_compare)
    incoming_df[norm_key] = incoming_df[incoming_right_key].apply(normalize_value_for_compare)

    merged = base_gdf.merge(
        incoming_df,
        on=norm_key,
        how="left",
        suffixes=("", "_incoming"),
    )

    match_mask = merged[norm_key].isin(incoming_df[norm_key].dropna())

    incoming_cols = [c for c in incoming_df.columns if c != incoming_right_key]
    matched_incoming_columns = set(normalized_matches.keys())

    for col in incoming_cols:
        base_has_same_name = col in base_gdf.columns
        incoming_name = f"{col}_incoming" if base_has_same_name else col

        if incoming_name not in merged.columns:
            continue

        target_col = normalized_matches.get(col)

        if target_col == geometry_name or col == geometry_name:
            merged.drop(columns=[incoming_name], inplace=True, errors="ignore")
            continue

        incoming_series = merged[incoming_name]

        if target_col:
            merged.loc[match_mask, target_col] = incoming_series[match_mask]
            if incoming_name != target_col:
                merged.drop(columns=[incoming_name], inplace=True, errors="ignore")
        else:
            if base_has_same_name and incoming_name.endswith("_incoming"):
                merged.rename(columns={incoming_name: col}, inplace=True)

    # Remove any stray right-key column copy
    if incoming_right_key in merged.columns and incoming_right_key != left_key:
        merged.drop(columns=[incoming_right_key], inplace=True)

    # Ensure all incoming-only columns exist even if the merge produced no _incoming column
    for col in incoming_cols:
        if col == geometry_name:
            continue
        if col in matched_incoming_columns:
            continue
        if col not in merged.columns:
            try:
                mapping = incoming_df.set_index(norm_key)[col].to_dict()
                merged[col] = merged[norm_key].map(mapping)
                try:
                    merged[col] = merged[col].astype(incoming_df[col].dtype)
                except Exception:
                    pass
            except Exception:
                merged[col] = pd.NA

    # Drop any leftover *_incoming columns just in case
    incoming_suffix_cols = [c for c in merged.columns if c.endswith("_incoming")]
    if incoming_suffix_cols:
        merged.drop(columns=incoming_suffix_cols, inplace=True, errors="ignore")

    # Ensure no duplicate or near-duplicate columns remain
    normalized_seen = {}
    columns_to_drop = []
    for col in merged.columns:
        if col == geometry_name:
            continue
        norm = normalize_for_compare(col)
        if norm in normalized_seen:
            columns_to_drop.append(col)
        else:
            normalized_seen[norm] = col
    if columns_to_drop:
        merged.drop(columns=columns_to_drop, inplace=True, errors="ignore")

    if norm_key in merged.columns:
        merged.drop(columns=[norm_key], inplace=True, errors="ignore")

    for col in merged.columns:
        if col == geometry_name:
            continue
        merged[col] = ensure_valid_gpkg_dtypes(merged[col])

    result = gpd.GeoDataFrame(merged, geometry=geometry_name, crs=base_gdf.crs)
    return sanitize_gdf_for_gpkg(result)


def read_pairs_from_zip(uploaded_zip):
    """Return list of datasets extracted from an uploaded ZIP archive."""
    dataset_pairs = []
    with tempfile.TemporaryDirectory() as tmpdir:
        zip_path = os.path.join(tmpdir, uploaded_zip.name)
        with open(zip_path, "wb") as tmp_zip:
            tmp_zip.write(uploaded_zip.getbuffer())

        with zipfile.ZipFile(zip_path, "r") as zf:
            _safe_extract_zip(zf, tmpdir)

        paired_files = {}
        for root, _, files in os.walk(tmpdir):
            for file in files:
                base, ext = os.path.splitext(file)
                ext_lower = ext.lower()
                full_path = os.path.join(root, file)
                if ext_lower == ".gpkg":
                    paired_files.setdefault(base, {})["gpkg"] = full_path
                elif ext_lower in [".csv", ".xlsx"]:
                    paired_files.setdefault(base, {})["data"] = full_path

        for base_name, paths in paired_files.items():
            if "gpkg" not in paths or "data" not in paths:
                continue

            gdf = gpd.read_file(paths["gpkg"])
            data_path = paths["data"]
            try:
                df = read_tabular_data(data_path)
                df = clean_empty_rows(df)
            except Exception:
                continue

            dataset_pairs.append(
                {
                    "base": base_name,
                    "gdf": gdf,
                    "df": df,
                    "source_zip": uploaded_zip.name,
                }
            )

    return dataset_pairs


if not gpkg_files:
    st.info("Upload at least one GeoPackage to begin merging.")
else:
    for gpkg_file in gpkg_files:
        gpkg_id = Path(gpkg_file.name).stem or "dataset"
        paste_df_key = f"df_from_paste_{gpkg_id}"
        st.markdown(f"<h4>{html.escape(gpkg_file.name)}</h4>", unsafe_allow_html=True)

        _reset_stream(gpkg_file)
        try:
            gdf = gpd.read_file(gpkg_file)
        except Exception as exc:
            st.error(f"Unable to read {gpkg_file.name}: {exc}")
            continue

        st.caption(f"Loaded {len(gdf):,} feature(s)")
        st.dataframe(gdf.head(PREVIEW_ROW_COUNT))

        st.markdown("**Pasted Data Input**")
        paste_text = st.text_area(
            f"Paste data for {gpkg_file.name}",
            height=150,
            placeholder="Paste tabular data for this GeoPackage...",
            key=f"paste_{gpkg_id}",
        )

        if isinstance(paste_text, str) and paste_text.strip():
            parsed = None
            try:
                parsed = parse_pasted_tabular_text(paste_text)
            except Exception:
                parsed = None

            if isinstance(parsed, pd.DataFrame):
                edited = st.data_editor(
                    parsed,
                    num_rows="dynamic",
                    key=f"pasted_data_editor_{gpkg_id}",
                )
                if isinstance(edited, pd.DataFrame):
                    cleaned = clean_empty_rows(edited)
                    if not cleaned.dropna(how="all").empty:
                        st.session_state[paste_df_key] = cleaned
                        st.success("Pasted data parsed successfully.")
                    else:
                        st.warning("Pasted data appears empty after cleaning.")
            else:
                st.warning(
                    "Unable to parse pasted text as a table. Please ensure it's tabular (TSV/CSV) or paste directly from Excel cells."
                )
        else:
            if paste_df_key in st.session_state:
                del st.session_state[paste_df_key]

        pasted_df = st.session_state.get(paste_df_key)
        df_for_merge = None
        df_source_description = None
        if isinstance(pasted_df, pd.DataFrame) and not pasted_df.dropna(how="all").empty:
            df_for_merge = pasted_df
            df_source_description = "Pasted data"
        elif shared_df is not None:
            df_for_merge = shared_df
            df_source_description = shared_source_label or "Uploaded tabular data"

        st.markdown("**Tabular Data Preview**")
        if df_for_merge is not None:
            if df_source_description:
                st.caption(f"Source: {df_source_description}")
            st.dataframe(df_for_merge.head(PREVIEW_ROW_COUNT))
        else:
            st.warning(
                "No tabular data detected for this GeoPackage. Paste data above or provide a CSV/Excel/reference workbook."
            )

        left_key = None
        right_key = None
        if df_for_merge is not None:
            st.markdown('<div class="section-title">Select join fields</div>', unsafe_allow_html=True)
            left_key = st.selectbox(
                "Field in GeoPackage",
                gdf.columns,
                key=f"left_key_{gpkg_id}",
            )
            right_key = st.selectbox(
                "Field in Data File",
                df_for_merge.columns,
                key=f"right_key_{gpkg_id}",
            )

        auto_name = gpkg_id or "updated_clean"
        suggested_name = name_memory.get(auto_name, auto_name)
        output_name = (
            st.text_input(
                "Name for the updated GeoPackage (without extension)",
                value=suggested_name,
                key=f"output_name_{gpkg_id}",
                help="This will also be used for the GeoPackage layer name.",
            ).strip()
            or auto_name
        )
        layer_name = output_name.replace(" ", "_")

        if st.button(f"Merge This GPKG ({gpkg_file.name})", key=f"merge_btn_{gpkg_id}"):
            if df_for_merge is None:
                st.warning(
                    f"{gpkg_file.name}: Provide pasted data or load a CSV/Excel/reference workbook before merging."
                )
                continue
            if not left_key or not right_key:
                st.warning(f"{gpkg_file.name}: Please select both join fields.")
                continue

            try:
                df_to_merge = df_for_merge.copy() if hasattr(df_for_merge, "copy") else df_for_merge
            except Exception:
                df_to_merge = df_for_merge

            try:
                merged_gdf = merge_without_duplicates(gdf.copy(), df_to_merge, left_key, right_key)
                st.success(f"Attributes merged for {gpkg_file.name}.")
                st.dataframe(merged_gdf.head())

                with tempfile.NamedTemporaryFile(suffix=".gpkg", delete=False) as tmp:
                    temp_path = tmp.name

                try:
                    safe_gdf = sanitize_gdf_for_gpkg(merged_gdf)
                    try:
                        existing = name_memory.get(auto_name)
                        if output_name and output_name != existing and output_name != auto_name:
                            set_saved_name(auto_name, output_name, name_memory)
                    except Exception:
                        pass

                    safe_gdf.to_file(temp_path, driver="GPKG", layer=layer_name)
                    with open(temp_path, "rb") as updated:
                        data_bytes = updated.read()

                    st.download_button(
                        f"Download {output_name}.gpkg",
                        data=data_bytes,
                        file_name=f"{output_name}.gpkg",
                        mime="application/geopackage+sqlite3",
                        key=f"download_{gpkg_id}",
                    )
                finally:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
            except Exception as exc:
                st.error(f"Error while merging {gpkg_file.name}: {exc}")

# ------------------- GEOMETRY CONVERSION (POINT to POLYGON) -----------------
with st.container():
    st.markdown('<div class="section-box tertiary">', unsafe_allow_html=True)
    st.markdown(
        '<div class="section-title">Geometry Conversion (Points to Polygons)</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p class='section-subtext'>Upload a GeoPackage (.gpkg) or File Geodatabase (.gdb), select a point layer, and convert points into rectangular polygons (length × width in meters). Download the updated file in the same format.</p>",
        unsafe_allow_html=True,
    )

    pt_to_poly_file = st.file_uploader(
        "Upload GeoPackage (.gpkg) or File Geodatabase (.gdb)",
        type=["gpkg", "gdb"],
        key="point_to_polygon_gpkg",
        accept_multiple_files=False,
    )

    if pt_to_poly_file:
        temp_input_path = None
        try:
            # Determine file format and create temp file with appropriate extension
            file_suffix = Path(pt_to_poly_file.name).suffix.lower()
            temp_suffix = file_suffix if file_suffix in [".gpkg", ".gdb"] else ".gpkg"
            with tempfile.NamedTemporaryFile(suffix=temp_suffix, delete=False) as tmp_in:
                tmp_in.write(pt_to_poly_file.getbuffer())
                temp_input_path = tmp_in.name

            # List all layers in the file (works for both .gpkg and .gdb)
            layer_names = pyogrio.list_layers(temp_input_path)
            if hasattr(layer_names, "name"):
                layer_list = list(layer_names["name"])
            else:
                layer_list = [li[0] if isinstance(li, (list, tuple)) else str(li) for li in layer_names]

            if not layer_list:
                st.warning("No layers found in the GeoPackage.")
            else:
                selected_layer = st.selectbox(
                    "Select layer to convert",
                    layer_list,
                    key="ptpoly_layer_select",
                )

                # Load the selected layer
                gdf = gpd.read_file(temp_input_path, layer=selected_layer)

                geom_types = gdf.geom_type.dropna().unique().tolist()
                has_point = any("point" in str(gt).lower() for gt in geom_types)

                if not has_point:
                    st.warning("Selected layer does not contain Point geometries.")
                else:
                    st.markdown("**Conversion Parameters**")
                    length_m = st.number_input(
                        "Length (meters; north-south)",
                        min_value=0.01,
                        value=50.0,
                        step=0.1,
                        key="ptpoly_length",
                    )
                    width_m = st.number_input(
                        "Width (meters; east-west)",
                        min_value=0.01,
                        value=50.0,
                        step=0.1,
                        key="ptpoly_width",
                    )
                    rotation_deg = st.number_input(
                        "Rotation (degrees clockwise)",
                        min_value=-360.0,
                        max_value=360.0,
                        value=0.0,
                        step=0.5,
                        key="ptpoly_rotation",
                    )

                    preview_map = st.checkbox(
                        "Show preview map",
                        value=True,
                        key="ptpoly_preview",
                    )

                    if st.button("Convert Points to Polygons", key="ptpoly_convert_btn"):
                        orig_crs = gdf.crs
                        if orig_crs is None:
                            st.warning(
                                "Source layer has no CRS; assuming EPSG:4326 (lat/lon). Results may be inaccurate."
                            )
                            orig_crs = "EPSG:4326"
                            gdf = gdf.set_crs(orig_crs, allow_override=True)

                        try:
                            metric_crs = "EPSG:3857"
                            gdf_metric = gdf.to_crs(metric_crs)
                            polys = []
                            for _, row in gdf_metric.iterrows():
                                geom = row.geometry
                                if geom is None or geom.is_empty or not geom.geom_type.lower().endswith("point"):
                                    polys.append(None)
                                    continue
                                x = geom.x
                                y = geom.y
                                half_w = float(width_m) / 2.0
                                half_l = float(length_m) / 2.0
                                poly = box(x - half_w, y - half_l, x + half_w, y + half_l)
                                try:
                                    if float(rotation_deg) % 360.0 != 0.0:
                                        poly = rotate(poly, -float(rotation_deg), origin=(x, y), use_radians=False)
                                except Exception:
                                    pass
                                polys.append(poly)

                            poly_gdf = gdf.copy()
                            poly_gdf["geometry"] = gpd.GeoSeries(polys, index=poly_gdf.index, crs=metric_crs)
                            poly_gdf = poly_gdf.set_geometry("geometry").to_crs(orig_crs)
                            st.success("Rectangular polygons generated.")
                            st.dataframe(poly_gdf.head())

                            # Preview map
                            if preview_map:
                                try:
                                    poly_wgs = poly_gdf.to_crs("EPSG:4326")
                                    geojson = json.loads(poly_wgs.to_json())
                                    centroids = poly_wgs.geometry.centroid
                                    lon_mean = float(centroids.x.mean()) if len(centroids) > 0 else 0.0
                                    lat_mean = float(centroids.y.mean()) if len(centroids) > 0 else 0.0

                                    geo_layer = pdk.Layer(
                                        "GeoJsonLayer",
                                        data=geojson,
                                        pickable=True,
                                        stroked=True,
                                        filled=True,
                                        get_fill_color=[255, 165, 0, 140],
                                        get_line_color=[0, 0, 0],
                                        line_width_min_pixels=1,
                                    )

                                    deck = pdk.Deck(
                                        layers=[geo_layer],
                                        initial_view_state=pdk.ViewState(
                                            longitude=lon_mean,
                                            latitude=lat_mean,
                                            zoom=14,
                                            pitch=0,
                                        ),
                                    )
                                    st.pydeck_chart(deck)
                                except Exception as exc:
                                    st.warning(f"Preview unavailable: {exc}")

                            # Save converted layer, preserving input format
                            temp_output_path = None
                            output_driver = None
                            output_mime = None
                            
                            # Determine output format based on input
                            input_ext = Path(pt_to_poly_file.name).suffix.lower()
                            if input_ext == ".gdb":
                                output_driver = "FileGDB"
                                output_mime = "application/x-filegdb"
                                temp_output_dir = tempfile.mkdtemp()
                                temp_output_path = os.path.join(temp_output_dir, "converted.gdb")
                            else:
                                output_driver = "GPKG"
                                output_mime = "application/geopackage+sqlite3"
                                with tempfile.NamedTemporaryFile(suffix=".gpkg", delete=False) as tmp_out:
                                    temp_output_path = tmp_out.name
                            
                            try:
                                safe_poly = sanitize_gdf_for_gpkg(poly_gdf)
                                safe_poly.to_file(temp_output_path, driver=output_driver, layer=selected_layer)
                                
                                # For .gdb, zip the .gdb folder; for .gpkg, read file directly
                                if output_driver == "FileGDB":
                                    gdb_zip_path = tempfile.mktemp(suffix=".zip")
                                    try:
                                        with zipfile.ZipFile(gdb_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                                            for root, _, files in os.walk(temp_output_path):
                                                for file in files:
                                                    file_path = os.path.join(root, file)
                                                    arcname = os.path.relpath(file_path, os.path.dirname(temp_output_path))
                                                    zf.write(file_path, arcname)
                                        with open(gdb_zip_path, "rb") as zf_in:
                                            output_bytes = zf_in.read()
                                        output_filename = pt_to_poly_file.name.replace(".gdb", "") + ".gdb.zip"
                                    finally:
                                        if os.path.exists(gdb_zip_path):
                                            os.remove(gdb_zip_path)
                                else:
                                    with open(temp_output_path, "rb") as converted:
                                        output_bytes = converted.read()
                                    output_filename = pt_to_poly_file.name

                                st.download_button(
                                    f"Download {output_filename}",
                                    data=output_bytes,
                                    file_name=output_filename,
                                    mime=output_mime,
                                    key="download_ptpoly_gpkg",
                                )
                            except Exception as exc:
                                st.error(f"Failed to prepare output file for download: {exc}")
                            finally:
                                if output_driver == "FileGDB" and os.path.exists(temp_output_path):
                                    shutil.rmtree(temp_output_path, ignore_errors=True)
                                elif output_driver == "GPKG" and temp_output_path and os.path.exists(temp_output_path):
                                    os.remove(temp_output_path)
                        except Exception as exc:
                            st.error(f"Conversion failed: {exc}")

        except Exception as exc:
            st.error(f"Unable to read file: {exc}")
        finally:
            if temp_input_path and os.path.exists(temp_input_path):
                os.remove(temp_input_path)
    else:
        st.info("Upload a GeoPackage (.gpkg) or File Geodatabase (.gdb) to begin.")

    st.markdown('</div>', unsafe_allow_html=True)

# ------------------- GEOMETRY CONVERSION (POLYGON to POINT) -----------------
with st.container():
    st.markdown('<div class="section-box tertiary">', unsafe_allow_html=True)
    st.markdown(
        '<div class="section-title">Geometry Conversion (Polygons to Points)</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p class='section-subtext'>Upload one or more GeoPackages to convert all polygon features into centroid points while keeping every attribute intact.</p>",
        unsafe_allow_html=True,
    )

    polygon_conversion_files = st.file_uploader(
        "Upload GeoPackage (.gpkg) for centroid conversion",
        type=["gpkg"],
        key="polygon_to_point_gpkg",
        accept_multiple_files=True,
    )

    converted_packages = []

    if polygon_conversion_files:
        for polygon_conversion_file in polygon_conversion_files:
            st.markdown(f"**Processing:** {polygon_conversion_file.name}")
            conversion_gdf = None
            temp_input_path = None
            source_layer_name = None
            try:
                with tempfile.NamedTemporaryFile(suffix=".gpkg", delete=False) as tmp_in:
                    tmp_in.write(polygon_conversion_file.getbuffer())
                    temp_input_path = tmp_in.name

                conversion_gdf = gpd.read_file(temp_input_path)
                try:
                    layer_info = pyogrio.list_layers(temp_input_path)
                    if hasattr(layer_info, "__iter__"):
                        # DataFrame-like; prefer name column
                        if hasattr(layer_info, "name"):
                            names = list(layer_info["name"])
                        else:
                            names = [li[0] for li in layer_info] if layer_info else []
                        if names:
                            source_layer_name = names[0]
                except Exception:
                    source_layer_name = None
                st.success(
                    f"Loaded GeoPackage with {len(conversion_gdf):,} feature(s) ready for conversion."
                )
            except Exception as exc:
                conversion_gdf = None
                st.error(f"Unable to read {polygon_conversion_file.name}: {exc}")
            finally:
                if temp_input_path and os.path.exists(temp_input_path):
                    os.remove(temp_input_path)

            if conversion_gdf is None:
                continue

            geom_types_raw = conversion_gdf.geom_type.dropna().unique().tolist()
            geom_types_clean = sorted({str(gt) for gt in geom_types_raw if str(gt).strip()})
            geom_types_display = ", ".join(geom_types_clean) if geom_types_clean else "Unknown"
            st.markdown(f"Detected geometry types: {geom_types_display}")

            has_polygon_geometry = any(
                "polygon" in str(geom_type).lower()
                for geom_type in geom_types_raw
            )
            has_point_geometry = any(
                "point" in str(geom_type).lower()
                for geom_type in geom_types_raw
            )

            target_layer_name = derive_layer_name_from_filename(
                polygon_conversion_file.name
            )

            if not has_polygon_geometry:
                if (
                    has_point_geometry
                    and source_layer_name
                    and source_layer_name != target_layer_name
                ):
                    st.info(
                        "No polygons detected, but updating the layer name to match the GeoPackage filename."
                    )
                    temp_output_path = None
                    try:
                        safe_copy = sanitize_gdf_for_gpkg(conversion_gdf)
                        with tempfile.NamedTemporaryFile(suffix=".gpkg", delete=False) as tmp_out:
                            temp_output_path = tmp_out.name
                        safe_copy.to_file(
                            temp_output_path,
                            driver="GPKG",
                            layer=target_layer_name,
                        )
                        with open(temp_output_path, "rb") as converted:
                            converted_packages.append(
                                (polygon_conversion_file.name, converted.read())
                            )
                        st.success(
                            f"Layer renamed to '{target_layer_name}' for {polygon_conversion_file.name}."
                        )
                    except Exception as exc:
                        st.error(
                            f"Failed to update layer name for {polygon_conversion_file.name}: {exc}"
                        )
                    finally:
                        if temp_output_path and os.path.exists(temp_output_path):
                            os.remove(temp_output_path)
                else:
                    st.info(
                        "This GeoPackage does not contain Polygon or MultiPolygon geometries, so no centroid conversion was performed."
                    )
                continue

            try:
                points_gdf = conversion_gdf.copy()
                points_gdf["geometry"] = conversion_gdf.geometry.centroid
                st.success("Centroid points generated for all polygon features.")
                st.dataframe(points_gdf.head())

                temp_output_path = None
                safe_points = sanitize_gdf_for_gpkg(points_gdf)
                try:
                    with tempfile.NamedTemporaryFile(suffix=".gpkg", delete=False) as tmp_out:
                        temp_output_path = tmp_out.name
                    safe_points.to_file(
                        temp_output_path,
                        driver="GPKG",
                        layer=target_layer_name,
                    )
                    with open(temp_output_path, "rb") as converted:
                        converted_packages.append(
                            (polygon_conversion_file.name, converted.read())
                        )
                except Exception as exc:
                    st.error(
                        f"Failed to prepare centroid GeoPackage for {polygon_conversion_file.name}: {exc}"
                    )
                finally:
                    if temp_output_path and os.path.exists(temp_output_path):
                        os.remove(temp_output_path)
            except Exception as exc:
                st.error(f"Failed to generate centroids for {polygon_conversion_file.name}: {exc}")
    else:
        st.info("Upload at least one GeoPackage to begin conversion.")

    if converted_packages:
        zip_bytes = None
        temp_zip_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as tmp_zip:
                temp_zip_path = tmp_zip.name
            with zipfile.ZipFile(temp_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for file_name, contents in converted_packages:
                    zf.writestr(file_name, contents)
            with open(temp_zip_path, "rb") as zip_file:
                zip_bytes = zip_file.read()
        except Exception as exc:
            zip_bytes = None
            st.error(f"Failed to package centroid GeoPackages: {exc}")
        finally:
            if temp_zip_path and os.path.exists(temp_zip_path):
                os.remove(temp_zip_path)

        if zip_bytes:
            st.download_button(
                "Download centroid GeoPackages (ZIP)",
                data=zip_bytes,
                file_name="centroid_points.zip",
                mime="application/zip",
            )

    st.markdown('</div>', unsafe_allow_html=True)

# --------------------- ZIP BUNDLE WORKFLOW --------------------------------
with st.container():
    st.markdown('<div class="section-box tertiary">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Batch ZIP Processing</div>', unsafe_allow_html=True)
    st.markdown(
        "<p class='section-subtext'>Process multiple GeoPackage + spreadsheet pairs by uploading ZIP archives that contain matching filenames.</p>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "Upload one or more ZIP archives. Each ZIP should contain a GeoPackage and a matching "
        "CSV/Excel file that share the same base name (e.g., `roads.gpkg` + `roads.xlsx`)."
    )

    uploaded_zips = st.file_uploader(
        "Upload zipped GeoPackage + spreadsheet bundles",
        type=["zip"],
        accept_multiple_files=True,
    )

    all_datasets = []
    if uploaded_zips:
        for uploaded_zip in uploaded_zips:
            try:
                all_datasets.extend(read_pairs_from_zip(uploaded_zip))
            except zipfile.BadZipFile:
                st.error(f"{uploaded_zip.name} is not a valid ZIP archive.")
            except Exception as exc:
                st.error(f"Failed to read {uploaded_zip.name}: {exc}")

    if not uploaded_zips:
        st.info("Start by uploading at least one ZIP file.")
    elif not all_datasets:
        st.warning("No valid GeoPackage + spreadsheet pairs were found in the uploaded ZIPs.")
    else:
        st.success(f"Loaded {len(all_datasets)} dataset(s) from the uploaded ZIP files.")
        st.write(
            "Configure the join fields and output names for each dataset, then click "
            "`Merge All Bundles` to generate the updated GeoPackages."
        )

        for idx, dataset in enumerate(all_datasets):
            with st.expander(
                f"Dataset {idx + 1}: {dataset['base']} ({dataset['source_zip']})",
                expanded=True,
            ):
                st.write("Select join fields for this dataset:")
                st.selectbox(
                    "Field in GeoPackage",
                    dataset["gdf"].columns,
                    key=f"left_key_{idx}",
                )
                st.selectbox(
                    "Field in Spreadsheet",
                    dataset["df"].columns,
                    key=f"right_key_{idx}",
                )
                ds_equipment_type = dataset["base"]
                ds_auto = f"{dataset['base']}_updated"
                ds_suggested = name_memory.get(ds_equipment_type, ds_auto)
                st.text_input(
                    "Output file name (without extension)",
                    value=ds_suggested,
                    key=f"output_name_{idx}",
                )

        if st.button("Merge All Bundles"):
            for idx, dataset in enumerate(all_datasets):
                left_key = st.session_state.get(f"left_key_{idx}")
                right_key = st.session_state.get(f"right_key_{idx}")
                output_name = (
                    st.session_state.get(f"output_name_{idx}", "").strip()
                    or f"{dataset['base']}_updated"
                )
                layer_name = output_name.replace(" ", "_")

                if not left_key or not right_key:
                    st.warning(
                        f"Dataset {dataset['base']} skipped: please select both join fields."
                    )
                    continue

                try:
                    merged_gdf = merge_without_duplicates(
                        dataset["gdf"], dataset["df"], left_key, right_key
                    )

                    with tempfile.NamedTemporaryFile(suffix=".gpkg", delete=False) as tmp:
                        temp_path = tmp.name

                    try:
                        safe_gdf = sanitize_gdf_for_gpkg(merged_gdf)
                        try:
                            ds_equipment_type = dataset["base"]
                            existing = name_memory.get(ds_equipment_type)
                            ds_auto = f"{dataset['base']}_updated"
                            if output_name and output_name != existing and output_name != ds_auto:
                                set_saved_name(ds_equipment_type, output_name, name_memory)
                        except Exception:
                            pass

                        safe_gdf.to_file(temp_path, driver="GPKG", layer=layer_name)
                        with open(temp_path, "rb") as updated:
                            data_bytes = updated.read()

                        st.success(f"{output_name}.gpkg is ready")
                        st.dataframe(merged_gdf.head())
                        st.download_button(
                            f"Download {output_name}.gpkg",
                            data=data_bytes,
                            file_name=f"{output_name}.gpkg",
                            mime="application/geopackage+sqlite3",
                        )
                    finally:
                        if os.path.exists(temp_path):
                            os.remove(temp_path)
                except Exception as exc:
                    st.error(f"Failed to merge dataset {dataset['base']}: {exc}")
    st.markdown('</div>', unsafe_allow_html=True)

with st.container():
    st.markdown('<div class="section-box tertiary">', unsafe_allow_html=True)
    st.markdown(
        '<div class="section-title">GeoPackage to File Geodatabase Conversion</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p class='section-subtext'>Translate GeoPackage datasets into ESRI File Geodatabases "
        "while preserving feature geometry, CRS definitions, and attribute schema.</p>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "Upload one or more GeoPackages to create engineering-grade File Geodatabase "
        "deliverables. Each layer name is aligned with the source filename, and the "
        "resulting .gdb folders are zipped for streamlined distribution.",
    )

    gpkg_conversion_files = st.file_uploader(
        "Upload GeoPackage datasets for File Geodatabase export",
        type=["gpkg"],
        accept_multiple_files=True,
    )

    converted_gdb_archives = []
    if not gpkg_conversion_files:
        st.info("Provide at least one GeoPackage to begin the conversion workflow.")
    else:
        for gpkg_file in gpkg_conversion_files:
            base_name = Path(gpkg_file.name).stem
            layer_name = base_name.replace(" ", "_")
            temp_gdb_root = tempfile.mkdtemp()
            gdb_directory = os.path.join(temp_gdb_root, f"{layer_name}.gdb")
            try:
                _reset_stream(gpkg_file)
                source_gdf = gpd.read_file(gpkg_file)
                safe_gdf = sanitize_gdf_for_gpkg(source_gdf)
                safe_gdf.to_file(gdb_directory, driver="FileGDB", layer=layer_name)

                archive_base = os.path.join(temp_gdb_root, f"{layer_name}.gdb")
                archive_path = shutil.make_archive(
                    archive_base,
                    "zip",
                    root_dir=temp_gdb_root,
                    base_dir=f"{layer_name}.gdb",
                )
                with open(archive_path, "rb") as archive_fh:
                    archive_bytes = archive_fh.read()

                converted_gdb_archives.append(
                    {
                        "file_name": f"{layer_name}.gdb.zip",
                        "data": archive_bytes,
                    }
                )

                feature_count = len(safe_gdf)
                crs_text = safe_gdf.crs.to_string() if safe_gdf.crs else "Unspecified CRS"
                st.success(
                    f"{gpkg_file.name} converted to File Geodatabase dataset '{layer_name}' "
                    f"({feature_count} spatial features, CRS: {crs_text})."
                )
                st.download_button(
                    f"Download {layer_name}.gdb.zip",
                    data=archive_bytes,
                    file_name=f"{layer_name}.gdb.zip",
                    mime="application/zip",
                    key=f"download_gdb_{gpkg_file.name}",
                )
            except Exception as exc:
                st.error(f"Failed to convert {gpkg_file.name}: {exc}")
            finally:
                shutil.rmtree(temp_gdb_root, ignore_errors=True)

    if converted_gdb_archives:
        st.success(
            f"Prepared {len(converted_gdb_archives)} File Geodatabase package(s) ready for delivery."
        )
        combined_zip_bytes = None
        combined_zip_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as tmp_bundle:
                combined_zip_path = tmp_bundle.name
            with zipfile.ZipFile(
                combined_zip_path, "w", compression=zipfile.ZIP_DEFLATED
            ) as bundle_zip:
                for archive in converted_gdb_archives:
                    bundle_zip.writestr(archive["file_name"], archive["data"])
            with open(combined_zip_path, "rb") as bundle_file:
                combined_zip_bytes = bundle_file.read()
        except Exception as exc:
            st.error(f"Failed to assemble the combined File Geodatabase ZIP: {exc}")
        finally:
            if combined_zip_path and os.path.exists(combined_zip_path):
                os.remove(combined_zip_path)

        if combined_zip_bytes:
            st.download_button(
                "Download all File Geodatabases (ZIP)",
                data=combined_zip_bytes,
                file_name="file_geodatabases.zip",
                mime="application/zip",
            )
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown(
    """
</div>

<div class='custom-footer'>Developed by Eng. IRANZI Prince Jean Claude</div>
""",
    unsafe_allow_html=True,
)
